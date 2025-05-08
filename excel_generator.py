# excel_generator.py
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define a gray fill for empty/placeholder cells
PLACEHOLDER_FILL_COLOR = "D3D3D3" # LightGray

def _apply_styles(ws, header_fill_color="000000", header_font_color="FFFFFF", default_row_height=30): # Added default_row_height
    """Applies common styling to a worksheet."""
    header_font = Font(color=header_font_color, bold=True)
    header_fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
    
    # Point 2: Vertical align middle all cells (default for data)
    general_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True) 
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True) # Explicit for header

    thin_border_side = Side(style='thin')
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    for row_idx, row_obj in enumerate(ws.iter_rows()): 
        is_header_row = (row_idx == 0)
        # Point 1: Make cells bigger so text can breathe a bit more (applied to data rows)
        if not is_header_row and default_row_height:
             ws.row_dimensions[row_obj[0].row].height = default_row_height 

        for cell_idx, cell in enumerate(row_obj):
            cell.border = cell_border
            if is_header_row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment 
            else: # Data rows
                cell.alignment = general_alignment 
                # Point 3: Horizontal and vertical align the number inside "Version #" field.
                # Assuming "Version #" is always the first column (index 0 for cell_idx)
                if cell_idx == 0: 
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Adjust column widths (simple auto-fit based on header or sample content)
    for col_idx, column_cells in enumerate(ws.columns):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_idx + 1)
        
        if ws.cell(row=1, column=col_idx + 1).value:
            max_length = len(str(ws.cell(row=1, column=col_idx + 1).value))

        for i in range(2, min(6, ws.max_row + 1)): # Check a few data rows (start from row 2)
             cell_value = ws.cell(row=i, column=col_idx + 1).value
             if cell_value:
                cell_len = len(str(cell_value).split('\n')[0]) 
                max_length = max(max_length, cell_len)
        
        adjusted_width = (max_length + 8) # Increased padding for breathing room
        header_val_for_width = ws.cell(row=1, column=col_idx + 1).value
        if header_val_for_width and str(header_val_for_width).lower() in ["body", "introductory text", "primary text", "link description", "text", "headline", "subject line"]:
             adjusted_width = max(adjusted_width, 50) 
        ws.column_dimensions[column_letter].width = min(adjusted_width, 70)


def create_excel_workbook(all_content_data, scraped_info, company_name_for_file):
    """Creates an Excel workbook with all generated content and styling."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet

    # --- Email Sheet ---
    if "email" in all_content_data and all_content_data["email"]:
        ws_email = wb.create_sheet(title="Email")
        email_headers = ["Version #", "Objective", "Headline", "Subject Line", "Body", "CTA"]
        ws_email.append(email_headers)
        for item in all_content_data["email"]:
            ws_email.append([
                item.get("Version #"), item.get("Objective"), item.get("Headline"),
                item.get("SubjectLine"), item.get("Body"), item.get("CTA")
            ])
        _apply_styles(ws_email, default_row_height=45) # Increased row height for emails

    # --- LinkedIn Sheet ---
    if "linkedin" in all_content_data and all_content_data["linkedin"]:
        ws_linkedin = wb.create_sheet(title="LinkedIn")
        linkedin_headers = ["Version #", "Ad Name", "Objective", "Introductory Text", "Image Copy", "Headline", "Destination", "CTA Button"]
        ws_linkedin.append(linkedin_headers)
        for item in all_content_data["linkedin"]:
            ws_linkedin.append([
                item.get("Version #"), item.get("AdName"), item.get("Objective"), item.get("IntroductoryText"),
                item.get("ImageCopy"), item.get("Headline"), item.get("Destination"), item.get("CTAButton")
            ])
        _apply_styles(ws_linkedin, default_row_height=45) # Increased row height

    # --- Facebook Sheet ---
    if "facebook" in all_content_data and all_content_data["facebook"]:
        ws_facebook = wb.create_sheet(title="Facebook")
        facebook_headers = ["Version #", "Ad Name", "Objective", "Primary Text", "Image Copy", "Headline", "Link Description", "Destination", "CTA Button"]
        ws_facebook.append(facebook_headers)
        for item in all_content_data["facebook"]:
            ws_facebook.append([
                item.get("Version #"), item.get("AdName"), item.get("Objective"), item.get("PrimaryText"),
                item.get("ImageCopy"), item.get("Headline"), item.get("LinkDescription"),
                item.get("Destination"), item.get("CTAButton")
            ])
        _apply_styles(ws_facebook, default_row_height=45) # Increased row height

    # --- Common styling for Google Ad Sheets ---
    google_header_font = Font(color="FFFFFF", bold=True)
    google_header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    google_header_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True) # Section headers left aligned
    google_data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border_side = Side(style='thin')
    google_cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    placeholder_fill_pattern = PatternFill(start_color=PLACEHOLDER_FILL_COLOR, end_color=PLACEHOLDER_FILL_COLOR, fill_type="solid")

    # --- Google Search Sheet --- Point 6
    if "google_search" in all_content_data and all_content_data["google_search"]:
        gs_data = all_content_data["google_search"]
        ws_gs = wb.create_sheet(title="Google Search")
        
        current_row = 1
        # Headlines Section
        ws_gs.row_dimensions[current_row].height = 25
        cell_h_title = ws_gs.cell(row=current_row, column=1, value="Headlines (15 total, Max 30 characters each)")
        cell_h_title.font = google_header_font
        cell_h_title.fill = google_header_fill
        cell_h_title.alignment = google_header_alignment
        cell_h_title.border = google_cell_border
        current_row += 1

        headlines = gs_data.get("headlines", [])
        for i in range(15): # Expect 15 headlines
            ws_gs.row_dimensions[current_row].height = 30 
            text_content = headlines[i] if i < len(headlines) else "" # Empty if not enough generated
            
            cell = ws_gs.cell(row=current_row, column=1, value=text_content)
            cell.alignment = google_data_alignment
            cell.border = google_cell_border
            if not text_content or "Placeholder" in text_content or "Error" in text_content:
                cell.fill = placeholder_fill_pattern # Mark empty or placeholder cells with gray
            current_row += 1
        
        current_row += 1 # Add a spacer row

        # Descriptions Section
        ws_gs.row_dimensions[current_row].height = 25
        cell_d_title = ws_gs.cell(row=current_row, column=1, value="Descriptions (4 total, Max 90 characters each)")
        cell_d_title.font = google_header_font
        cell_d_title.fill = google_header_fill
        cell_d_title.alignment = google_header_alignment
        cell_d_title.border = google_cell_border
        current_row += 1

        descriptions = gs_data.get("descriptions", [])
        for i in range(4): # Expect 4 descriptions
            ws_gs.row_dimensions[current_row].height = 45 
            text_content = descriptions[i] if i < len(descriptions) else "" # Empty if not enough generated

            cell = ws_gs.cell(row=current_row, column=1, value=text_content)
            cell.alignment = google_data_alignment
            cell.border = google_cell_border
            if not text_content or "Placeholder" in text_content or "Error" in text_content:
                cell.fill = placeholder_fill_pattern # Mark empty or placeholder cells with gray
            current_row += 1
        
        ws_gs.column_dimensions['A'].width = 100 

    # --- Google Display Sheet --- Point 6
    if "google_display" in all_content_data and all_content_data["google_display"]:
        gd_data = all_content_data["google_display"]
        ws_gd = wb.create_sheet(title="Google Display")

        current_row = 1
        # Headlines Section
        ws_gd.row_dimensions[current_row].height = 25
        cell_h_title_gd = ws_gd.cell(row=current_row, column=1, value="Headlines (5 total, Max 30 characters each)")
        cell_h_title_gd.font = google_header_font
        cell_h_title_gd.fill = google_header_fill
        cell_h_title_gd.alignment = google_header_alignment
        cell_h_title_gd.border = google_cell_border
        current_row += 1

        headlines_gd = gd_data.get("headlines", [])
        for i in range(5): # Expect 5 headlines
            ws_gd.row_dimensions[current_row].height = 30
            text_content = headlines_gd[i] if i < len(headlines_gd) else "" # Empty if not enough generated
            
            cell = ws_gd.cell(row=current_row, column=1, value=text_content)
            cell.alignment = google_data_alignment
            cell.border = google_cell_border
            if not text_content or "Placeholder" in text_content or "Error" in text_content:
                cell.fill = placeholder_fill_pattern # Mark empty or placeholder cells with gray
            current_row += 1

        current_row += 1 # Add a spacer row

        # Descriptions Section
        ws_gd.row_dimensions[current_row].height = 25
        cell_d_title_gd = ws_gd.cell(row=current_row, column=1, value="Descriptions (5 total, Max 90 characters each)")
        cell_d_title_gd.font = google_header_font
        cell_d_title_gd.fill = google_header_fill
        cell_d_title_gd.alignment = google_header_alignment
        cell_d_title_gd.border = google_cell_border
        current_row += 1

        descriptions_gd = gd_data.get("descriptions", [])
        for i in range(5): # Expect 5 descriptions
            ws_gd.row_dimensions[current_row].height = 45
            text_content = descriptions_gd[i] if i < len(descriptions_gd) else "" # Empty if not enough generated

            cell = ws_gd.cell(row=current_row, column=1, value=text_content)
            cell.alignment = google_data_alignment
            cell.border = google_cell_border
            if not text_content or "Placeholder" in text_content or "Error" in text_content:
                cell.fill = placeholder_fill_pattern # Mark empty or placeholder cells with gray
            current_row += 1
        
        ws_gd.column_dimensions['A'].width = 100

    # --- Reasoning Sheet ---
    ws_reasoning = wb.create_sheet(title="Reasoning")
    ws_reasoning.append(["Scraped Client Data & Generation Reasoning"])
    ws_reasoning.merge_cells('A1:B1') 
    
    title_cell = ws_reasoning['A1']
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    current_row_reasoning = 2 
    scraped_data_map = {
        "Company Name": scraped_info.get('company_name'),
        "Tagline": scraped_info.get('tagline'),
        "Mission Statement": scraped_info.get('mission_statement'),
        "Industry": scraped_info.get('industry'),
        "Products/Services": ", ".join(scraped_info.get('products_services', [])) if scraped_info.get('products_services') else 'N/A',
        "USPs/Value Proposition": scraped_info.get('usps_value_proposition'),
        "Target Audience": scraped_info.get('target_audience'),
        "Tone of Voice": scraped_info.get('tone_of_voice'),
        "CTAs from Website": ", ".join(scraped_info.get('ctas', [])) if scraped_info.get('ctas') else 'N/A'
    }

    for key, value in scraped_data_map.items():
        ws_reasoning.row_dimensions[current_row_reasoning].height = 35 # Increased height
        ws_reasoning.cell(row=current_row_reasoning, column=1, value=key).font = Font(bold=True)
        ws_reasoning.cell(row=current_row_reasoning, column=2, value=str(value))
        for col in [1,2]:
            cell = ws_reasoning.cell(row=current_row_reasoning, column=col)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True) # Vertical center
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        current_row_reasoning += 1
    
    ws_reasoning.row_dimensions[current_row_reasoning].height = 15 # Spacer row height
    ws_reasoning.cell(row=current_row_reasoning, column=1, value="") 
    current_row_reasoning +=1

    ws_reasoning.row_dimensions[current_row_reasoning].height = 25
    ws_reasoning.cell(row=current_row_reasoning, column=1, value="Reasoning for Content Generation:").font = Font(bold=True, size=12)
    ws_reasoning.merge_cells(start_row=current_row_reasoning, start_column=1, end_row=current_row_reasoning, end_column=2)
    merged_reasoning_header = ws_reasoning.cell(row=current_row_reasoning, column=1)
    merged_reasoning_header.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    merged_reasoning_header.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    current_row_reasoning += 1
    
    reasoning_text_content = all_content_data.get("reasoning_text", "Reasoning not available.")
    reasoning_text_cell = ws_reasoning.cell(row=current_row_reasoning, column=1, value=reasoning_text_content)
    # Calculate required rows for reasoning text to make it breathe
    # Approx 70 chars per line for column B width, 2 lines per 30px height
    num_lines = sum([len(line) // 70 + 1 for line in reasoning_text_content.split('\n')])
    num_rows_for_reasoning = max(6, num_lines // 2 + 1) # at least 6 rows, or more if needed

    ws_reasoning.merge_cells(start_row=current_row_reasoning, start_column=1, end_row=current_row_reasoning + num_rows_for_reasoning -1 , end_column=2) 
    reasoning_text_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True) # Top align long text
    
    # Apply border to all cells in the merged range
    merged_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for r_m in range(current_row_reasoning, current_row_reasoning + num_rows_for_reasoning):
        ws_reasoning.row_dimensions[r_m].height = 30 # Set consistent height for reasoning rows
        for c_m in range(1, 3): # Columns A and B
            ws_reasoning.cell(row=r_m, column=c_m).border = merged_border


    ws_reasoning.column_dimensions['A'].width = 30
    ws_reasoning.column_dimensions['B'].width = 70

    # Save to a BytesIO object
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes